import pandas as pd
from openpyxl import load_workbook

df1 = pd.DataFrame([["4/1/22", "8",'5699:03:01','5699:03:01', '0']], columns=["Dag", "Tijdbestek", "Start Draai Uren", "Eind Draai Uren", "NettoDraaiUren"])

book = load_workbook('metingenFreesafdeling.xlsx')
writer = pd.ExcelWriter('metingenFreesafdeling.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}
sheetname = 'Hermle 201'

df1.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

writer.save()